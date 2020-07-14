import openpyxl
import xlwt 
from classes.data_miner import *
#TODO: create a funciton to recreate a new list in an excel -> if not found

"""_____________________________________GLOBAL SCOPE_____________________________________"""

#administration urls
yescom_url = "https://yescom.admin.s5.upgates.com/"
easy_print_url = "https://easy-print.admin.s5.upgates.com/"
nabitabaterka_url = "https://nabitabaterka.admin.s10.upgates.com/"
excel_url = "C:\\Users\\rober\\Dropbox\\interní\\Yescom analyza\\2016-2018_denni_obraty.xlsx"

turnover_total = []
package_cost = 75
order_number = 0

"""_______________________________MINING DATA ->data_miner_______________________________"""

def mine_data(**name_adress):
    """mine data from urls"""
    global order_number
    for admin, url in name_adress.items():
        if "_sk" in admin:
            exchange = 26
        else: #checking _sk version for CZ/EUR exchange
            exchange = 1
        admin = Data_miner() #init class data_miner
        admin.load_page(url) #mine data - miner function
        #reduce turonver by package costs
        turnover = (admin.list_name[0]*exchange)-(package_cost*admin.list_name[1])
        order_number += admin.list_name[1]
        turnover_total.append(turnover) 
    return order_number, turnover_total

"""__________________________________RECORD-DATA-EXCEL__________________________________"""

def record_to_excel():
    """writing turnover data to an excel"""
    #locate the correct sheet and open
    y = date.today()
    month_name = y.strftime("%m")+" "+y.strftime("%Y")
    day_row = int(y.strftime("%d"))+1 #calculate row number basen on a day
    turnover_excel = openpyxl.load_workbook(excel_url) #open an turnover excel on dropbox
    active_list = turnover_excel[month_name]
    #list has to be in correct format to be located! 01 2020 02 2020 etc.
    #change number_order - locate -> write
    number_orders = active_list.cell(row = day_row, column = 10)
    number_orders.value = order_number
    #number order is always only 1x
    print(order_number)
    print(turnover_total)
    for i in range(len(turnover_total)):
        #more administration more values - template shouled be changed then
        #but no need for change the code
        cell_change = active_list.cell(row = day_row, column = i+3)
        cell_change.value = turnover_total[i]
    #Save and exit an excell file
    turnover_excel.save(excel_url)

"""_______________________________________MAIN_______________________________________"""

def main():
    """init data_miner -> mine_data -> write_data"""
    mine_data(yescom_cz = yescom_url, 
    easy_print_sk = easy_print_url, 
    nabitabaterka_cz = nabitabaterka_url)
    record_to_excel()

if __name__ == "__main__":
    main()